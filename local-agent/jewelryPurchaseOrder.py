#!/usr/bin/env python3
"""
주얼리 발주서 생성기 (나비스트 등) — 제품 이미지 임베드 + 편집가능 엑셀.

기존 엑셀은 이미지 칸이 수식(#VALUE!)이라 깨졌음 → 실제 이미지를 셀에 임베드해 혼선 방지.
재사용: items JSON (productNo/refNo/name/qty/note/imageUrl) 넘기면 대표이미지 받아 발주서 생성.

사용:  python3 jewelryPurchaseOrder.py <items.json> <out.xlsx> [발주일] [납기]
items.json: [{"refNo":"PVJ028SV","name":"...","qty":80,"note":"","imagePath":"/tmp/x.jpg","sku":"P00000BU"}]
"""
import sys, json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

def build(items, out_path, order_date="", due=""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주서"

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="F2F0EB")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── 타이틀 ──
    ws.merge_cells("A1:H1")
    ws["A1"] = "발  주  서   ·   JEWELRY PURCHASE ORDER"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── 발주 정보 ──
    info = [
        ("발주처 (甲)", "해리엇와치스 (브랜드: 폴바이스) · 209-27-70599 · 홍성조"),
        ("소재지", "경기도 용인시 기흥구 중부대로 184 힉스유타워 717-2호 · harriotwatches@gmail.com"),
        ("수주처 (乙)", "(주)나비스트"),
        ("발주일 / 납기", f"{order_date or '____-__-__'}   /   {due or '협의 (리드타임 최소 3주)'}"),
    ]
    r = 2
    for k, v in info:
        ws.merge_cells(f"A{r}:B{r}"); ws.merge_cells(f"C{r}:H{r}")
        ws[f"A{r}"] = k; ws[f"A{r}"].font = Font(bold=True, size=10, color="666666")
        ws[f"A{r}"].alignment = left
        ws[f"C{r}"] = v; ws[f"C{r}"].alignment = left; ws[f"C{r}"].font = Font(size=10)
        r += 1
    r += 1

    # ── 표 헤더 ──
    headers = ["이미지", "SKU", "Ref. No.", "상품명", "발주수량", "단가", "금액", "비고(출고예정/단종)"]
    widths  = [20, 13, 12, 30, 11, 12, 14, 20]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, size=10); c.fill = hdr_fill
        c.alignment = center; c.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 22
    r += 1

    # ── 품목 행 (이미지 임베드) ──
    total_qty = 0
    for it in items:
        ws.row_dimensions[r].height = 92
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = border
        # 이미지
        if it.get("imagePath"):
            try:
                img = XLImage(it["imagePath"]); img.width = 118; img.height = 118
                ws.add_image(img, f"A{r}")
            except Exception:
                ws.cell(row=r, column=1, value="(이미지)")
        ws.cell(row=r, column=2, value=it.get("sku", "")).alignment = center
        ws.cell(row=r, column=3, value=it.get("refNo", "")).alignment = center
        ws.cell(row=r, column=4, value=it.get("name", "")).alignment = left
        qc = ws.cell(row=r, column=5, value=it.get("qty", "")); qc.alignment = center; qc.font = Font(bold=True)
        ws.cell(row=r, column=6, value="").alignment = center   # 단가(나비스트 기입)
        ws.cell(row=r, column=7, value="").alignment = center   # 금액
        ws.cell(row=r, column=8, value=it.get("note", "")).alignment = left
        total_qty += it.get("qty", 0) if isinstance(it.get("qty"), int) else 0
        r += 1

    # ── 합계 ──
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"] = "합계 수량"; ws[f"A{r}"].font = Font(bold=True); ws[f"A{r}"].alignment = center
    ws.cell(row=r, column=5, value=total_qty).font = Font(bold=True)
    ws.cell(row=r, column=5).alignment = center
    for col in range(1, 9):
        ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FAF9F6")
    r += 2

    # ── 비고 ──
    ws.merge_cells(f"A{r}:H{r}")
    ws[f"A{r}"] = "· 단가·금액은 나비스트 확정 견적 기준 기입 부탁드립니다.  · 여름 성수기 재입고 건 — 가능 시 일부 조기 분납 협의 희망."
    ws[f"A{r}"].font = Font(size=9, color="888888"); ws[f"A{r}"].alignment = left

    wb.save(out_path)
    print(f"저장: {out_path} ({len(items)}품목, 합계 {total_qty})")

if __name__ == "__main__":
    items = json.load(open(sys.argv[1], encoding="utf-8"))
    out = sys.argv[2]
    od = sys.argv[3] if len(sys.argv) > 3 else ""
    due = sys.argv[4] if len(sys.argv) > 4 else ""
    build(items, out, od, due)
