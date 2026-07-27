from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path("outputs/kakao_gift_product_proposal_20260616.xlsx")


summary_rows = [
    ["1차 업로드 추천", "악성 주얼리 처분 + 봄 오픈링 테스트", "12개 SKU 내외", "2주 테스트 후 클릭/찜/주문 없으면 할인율 상향"],
    ["재고 리스크", "현재 카카오 판매 추정 상품 중 품절", "BT/BU/BW", "신규 업로드보다 품절/대체 연결 먼저 처리"],
    ["반지 테스트", "모두 오픈링 전제", "GV/GX/HY 중심", "사이즈 리스크 낮으므로 봄 시즌 소량 테스트 적합"],
    ["가격 전략", "정가 대비 20~35% 카카오 전용 할인가", "메시지카드/단독포장 포함", "처분 목적 상품은 초기부터 할인 명확히"],
]


columns = [
    "구분",
    "우선순위",
    "SKU",
    "기존 상품명",
    "제안 상품명",
    "상품 유형",
    "현재 판매가",
    "카카오 제안가",
    "할인율",
    "현재고",
    "최근 90일 판매",
    "최근 30일 판매",
    "판매세트 구성 제안",
    "상품명/카피 방향",
    "운영 제안",
    "비고",
]


rows = [
    [
        "반지 테스트",
        1,
        "P00000GV",
        "네이쳐 플로우 반지",
        "사이즈 걱정 없는 봄 오픈링 선물 - 네이쳐 플로우",
        "오픈링/반지",
        78000,
        59000,
        "24%",
        13,
        0,
        0,
        "단품 + 메시지카드. GX/HY와 2종 택1 확장 가능",
        "봄, 자연스러운 곡선, 사이즈 걱정 없는 선물",
        "2주 소량 테스트. 반응 없으면 49,000원까지 인하",
        "반지 메인 테스트 1순위",
    ],
    [
        "반지 테스트",
        2,
        "P00000GX",
        "오가닉 실버 텍스쳐 반지",
        "봄날의 손끝 선물 - 오가닉 실버 오픈링",
        "오픈링/반지",
        96000,
        69000,
        "28%",
        7,
        0,
        0,
        "HY와 디자인/재고 중복 확인 후 옵션 통합",
        "실버, 오가닉 텍스처, 프리사이즈 오픈링",
        "소량 테스트. 상세에 오픈링 조절 가능 강조",
        "고가라 할인 명확히 필요",
    ],
    [
        "반지 테스트",
        3,
        "P00000HY",
        "오가닉 실버 텍스처드 링",
        "사이즈 부담 없는 실버 오픈링 선물",
        "오픈링/반지",
        96000,
        69000,
        "28%",
        7,
        0,
        0,
        "GX와 동일/유사 디자인이면 단일 상품 옵션으로 묶기",
        "심플한 실버 링, 데일리 오픈링",
        "GX와 중복이면 한 상품으로 합쳐 재고 소진",
        "상품명 중복 여부 확인 필요",
    ],
    [
        "반지 테스트",
        4,
        "P00000HZ",
        "데이지 플라워 링 - 실버",
        "봄꽃처럼 전하는 데이지 실버 오픈링",
        "오픈링/반지",
        50000,
        39000,
        "22%",
        2,
        2,
        0,
        "단품. 재고 2개라 테스트 노출 또는 재입고 판단용",
        "봄, 데이지, 가벼운 응원 선물",
        "재고 적음. 품절 후 반응 있으면 재입고 검토",
        "카카오 판매 추정 이력 있음",
    ],
    [
        "팔찌 처분",
        1,
        "P00000FJ",
        "큐빅 커프 팔찌 - 골드",
        "손목에 반짝이는 큐빅 커프 팔찌 선물 - 골드",
        "팔찌",
        59000,
        45000,
        "24%",
        49,
        1,
        0,
        "FS 실버와 2종 택1. 메시지카드 포함",
        "부담 없는 생일/감사 선물, 큐빅 포인트",
        "실버 커프와 옵션 묶음으로 빠른 소진",
        "현재고 높고 판매 거의 없음",
    ],
    [
        "팔찌 처분",
        2,
        "P00000FS",
        "큐빅 커프 팔찌 - 실버",
        "손목에 반짝이는 큐빅 커프 팔찌 선물 - 실버",
        "팔찌",
        59000,
        45000,
        "24%",
        43,
        5,
        2,
        "FJ 골드와 2종 택1. 메시지카드 포함",
        "카카오 테니스 팔찌 품절 대체 후보",
        "신규 업로드 1순위. 2주간 반응 체크",
        "약한 판매 신호 있음",
    ],
    [
        "팔찌 처분",
        3,
        "P00000DL",
        "C라인 에나멜 커프 팔찌 블루 - 실버",
        "컬러 포인트 C라인 에나멜 커프 팔찌 - 블루",
        "팔찌",
        49000,
        35000,
        "29%",
        38,
        1,
        0,
        "DN 레드와 2종 택1",
        "컬러 포인트, 가벼운 데일리 팔찌 선물",
        "처분 목적. 초기부터 30% 전후 할인",
        "현재고 높고 판매 거의 없음",
    ],
    [
        "팔찌 처분",
        4,
        "P00000DN",
        "C라인 에나멜 커프 팔찌 레드 - 로즈골드",
        "컬러 포인트 C라인 에나멜 커프 팔찌 - 레드",
        "팔찌",
        49000,
        35000,
        "29%",
        15,
        0,
        0,
        "DL 블루와 2종 택1",
        "로즈골드, 레드 포인트, 기분전환 선물",
        "DL과 한 상품 옵션으로 묶기",
        "재고 중간",
    ],
    [
        "귀걸이 처분",
        1,
        "P000000P",
        "(4mm) 큐빅 귀걸이 - 크리스탈 골드",
        "작고 반짝이는 4mm 큐빅 귀걸이 선물 - 크리스탈 골드",
        "귀걸이",
        32000,
        25000,
        "22%",
        41,
        0,
        0,
        "BH/BG/Q와 4종 택1",
        "부담 없는 데일리 귀걸이, 작은 선물",
        "카카오 저가 선물 라인으로 업로드",
        "현재고 많음",
    ],
    [
        "귀걸이 처분",
        2,
        "P00000BH",
        "(4mm) 큐빅 귀걸이 - 블랙",
        "작고 시크한 4mm 큐빅 귀걸이 선물 - 블랙",
        "귀걸이",
        34000,
        25000,
        "26%",
        39,
        0,
        0,
        "4mm 큐빅 귀걸이 4종 택1",
        "블랙 큐빅, 시크한 데일리 선물",
        "4종 옵션형으로 묶어 상세 제작 효율화",
        "면세점 출고 1개 반영",
    ],
    [
        "귀걸이 처분",
        3,
        "P000000Q",
        "(4mm) 큐빅 귀걸이 - 블랙 골드",
        "작고 시크한 4mm 큐빅 귀걸이 선물 - 블랙 골드",
        "귀걸이",
        32000,
        25000,
        "22%",
        20,
        0,
        0,
        "4mm 큐빅 귀걸이 4종 택1",
        "블랙 골드 포인트, 부담 없는 선물",
        "P/BH/BG와 묶어서 업로드",
        "판매 이력 거의 없음",
    ],
    [
        "귀걸이 처분",
        4,
        "P00000BL",
        "트리오 큐빅 하프 후프 귀걸이 - 골드",
        "트리오 큐빅 하프 후프 귀걸이 선물 - 골드",
        "귀걸이",
        30000,
        24000,
        "20%",
        33,
        0,
        0,
        "BK 실버와 2종 택1",
        "데일리 하프후프, 가벼운 생일 선물",
        "저가 귀걸이 라인으로 빠르게 테스트",
        "골드 재고 많음",
    ],
    [
        "목걸이 처분",
        1,
        "P00000IF",
        "화이트 오팔 원석 목걸이",
        "화이트 오팔 원석 목걸이 선물",
        "목걸이",
        80000,
        59000,
        "26%",
        48,
        0,
        0,
        "IB 로즈쿼츠/ID 터키석과 원석 목걸이 3종 택1",
        "원석, 응원, 의미 있는 선물",
        "원석 스토리형 상세로 처분 테스트",
        "현재고 높고 판매 없음",
    ],
    [
        "목걸이 처분",
        2,
        "P00000IB",
        "로즈쿼츠 원석 목걸이",
        "로즈쿼츠 원석 목걸이 선물",
        "목걸이",
        80000,
        59000,
        "26%",
        47,
        0,
        0,
        "IF/ID와 원석 목걸이 3종 택1",
        "로즈쿼츠, 마음을 전하는 선물",
        "화이트 오팔과 한 상품 옵션으로 묶기",
        "현재고 높고 판매 없음",
    ],
    [
        "목걸이 처분",
        3,
        "P00000FX",
        "8mm 크리스탈 진주 목걸이",
        "은은한 크리스탈 진주 목걸이 선물 - 8mm",
        "목걸이",
        90000,
        59000,
        "34%",
        29,
        0,
        0,
        "FW 6mm와 2종 택1",
        "진주, 단정한 데일리 선물",
        "가격 저항 예상. 30% 이상 할인 권장",
        "현재고 중상",
    ],
    [
        "목걸이 처분",
        4,
        "P00000CE",
        "볼 체인 목걸이 - 골드",
        "데일리 볼 체인 목걸이 선물 - 골드",
        "목걸이",
        41000,
        32000,
        "22%",
        29,
        0,
        0,
        "CD 실버와 2종 택1",
        "심플 체인, 데일리 목걸이 선물",
        "저가 목걸이 라인으로 업로드",
        "실버는 90일 3개 판매 신호 있음",
    ],
    [
        "보류/매핑",
        1,
        "P00000CC",
        "큐빅 펜던드 목걸이 - 실버",
        "스와로브스키 크리스탈 큐빅 펜던트 실버 목걸이",
        "목걸이",
        82000,
        59000,
        "28%",
        19,
        0,
        0,
        "단품. 기존 카카오 상품명과 SKU 매핑 확인",
        "이미 카카오에서 유사명으로 최근 30일 8개 발주",
        "신규 업로드 전 기존 상품 매핑/재고 차감 확인",
        "신규 후보라기보다 매핑 보정 대상",
    ],
]


notes_rows = [
    ["데이터 기준일", "2026-06-16"],
    ["재고 기준", "Supabase paulvice_inventory_v1 + Cafe24 최근 주문 + 카카오 발주서 보정"],
    ["카카오 매핑 이슈", "kakao_gift_sku_map이 0건이라 발주서 상품명으로 현재 카카오 판매 SKU를 추정"],
    ["악성 재고 기준", "현재고가 있고 최근 90일 판매가 0~1개인 주얼리 SKU"],
    ["반지 전제", "사용자 확인: 반지는 모두 오픈링. 사이즈 리스크 낮아 봄 테스트 가능"],
    ["운영 기간", "1차 테스트 2주. 클릭/찜/주문이 없으면 할인율 10%p 상향 또는 제외"],
]


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "추천 요약"
    ws.append(["항목", "내용", "대상", "운영 메모"])
    for row in summary_rows:
        ws.append(row)
    style_sheet(ws)
    set_widths(ws, [20, 34, 24, 52])
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 42

    detail = wb.create_sheet("업로드 후보 상세")
    detail.append(columns)
    for row in rows:
        detail.append(row)
    style_sheet(detail)
    set_widths(detail, [16, 10, 14, 32, 42, 16, 14, 14, 10, 10, 14, 14, 42, 42, 42, 36])
    for row in range(2, detail.max_row + 1):
        detail.row_dimensions[row].height = 66
    for col in ["G", "H"]:
        for cell in detail[col][1:]:
            cell.number_format = '#,##0"원"'
    for cell in detail["J"][1:] + detail["K"][1:] + detail["L"][1:]:
        cell.number_format = "0"

    fills = {
        "반지 테스트": "DCFCE7",
        "팔찌 처분": "FEF3C7",
        "귀걸이 처분": "DBEAFE",
        "목걸이 처분": "FCE7F3",
        "보류/매핑": "E5E7EB",
    }
    for row in range(2, detail.max_row + 1):
        fill = PatternFill("solid", fgColor=fills.get(detail.cell(row, 1).value, "FFFFFF"))
        for col in range(1, detail.max_column + 1):
            detail.cell(row, col).fill = fill

    notes = wb.create_sheet("운영 메모")
    notes.append(["항목", "내용"])
    for row in notes_rows:
        notes.append(row)
    style_sheet(notes)
    set_widths(notes, [24, 90])
    for row in range(2, notes.max_row + 1):
        notes.row_dimensions[row].height = 38

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    build()
    wb = load_workbook(OUTPUT, data_only=False)
    expected = {"추천 요약", "업로드 후보 상세", "운영 메모"}
    if set(wb.sheetnames) != expected:
        raise SystemExit(f"Unexpected sheets: {wb.sheetnames}")
    ws = wb["업로드 후보 상세"]
    if ws.max_row < 10 or ws.max_column != len(columns):
        raise SystemExit("Detail sheet verification failed")
    print(f"saved={OUTPUT}")
    print(f"detail_rows={ws.max_row - 1}")
